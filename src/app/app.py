# -*- coding: utf-8 -*-
"""
Договорной контур — веб-приложение.

Управленческий контур: боковое меню, «Мой день» с проверками,
доска стадий, реестр, карточка договора с разбором замечаний,
контрагенты, гарантии, пользователи и права.
"""

from __future__ import annotations

import os
import secrets
import threading
from functools import wraps

import hashlib
import uuid

from flask import (
    Flask, session, request, redirect, url_for, render_template, abort, flash, g,
    send_file
)
from werkzeug.security import check_password_hash, generate_password_hash

import db
import review_engine
import text_extract

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY") or secrets.token_hex(32)
app.config["MAX_CONTENT_LENGTH"] = 60 * 1024 * 1024  # до 60 МБ на файл

FILES_DIR = os.getenv("FILES_DIR") or "C:/Users/claude/dogovor/files"

FILE_KINDS = [
    ("contract", "Договор"), ("estimate", "Смета"), ("ks2", "КС-2"),
    ("ks3", "КС-3"), ("act", "Акт"), ("supplement", "Допсоглашение"),
    ("protocol", "Протокол разногласий"), ("scan", "Скан"), ("other", "Другое"),
]
FILE_KIND_RU = dict(FILE_KINDS)

STAGE_RU = {
    "draft": "черновик", "internal_review": "согласование",
    "legal_review": "проверка юриста", "at_counterparty": "у заказчика",
    "in_progress": "в работе", "completed": "завершён", "warranty": "на гарантии",
    "archived": "архив", "cancelled": "аннулирован", "on_hold": "на стопе",
}
# порядок стадий на доске
STAGE_ORDER = ["draft", "internal_review", "legal_review", "at_counterparty",
               "in_progress", "completed", "on_hold", "cancelled"]
SEV_RU = {"critical": "критично", "warning": "предупреждение", "info": "к сведению"}
KIND_RU = {"commercial": "коммерческий", "budget": "бюджет", "uk_tsj": "УК/ТСЖ",
           "government": "госконтракт", "individual": "физлицо/ИП"}

def human_size(n):
    if n is None:
        return "—"
    for unit in ("Б", "КБ", "МБ"):
        if n < 1024:
            return f"{n:.0f} {unit}" if unit == "Б" else f"{n:.1f} {unit}"
        n /= 1024
    return f"{n:.1f} ГБ"


from markupsafe import Markup

_ICONS = {
    "home": '<path d="M3 11l9-8 9 8"/><path d="M5 10v10h14V10"/>',
    "board": '<rect x="3" y="4" width="5" height="16" rx="1"/><rect x="10" y="4" width="5" height="11" rx="1"/><rect x="17" y="4" width="4" height="16" rx="1"/>',
    "clock": '<circle cx="12" cy="12" r="9"/><path d="M12 7v5l3 2"/>',
    "task": '<rect x="3" y="3" width="18" height="18" rx="3"/><path d="M8 12l3 3 5-6"/>',
    "shield": '<path d="M12 3l7 3v6c0 4-3 7-7 9-4-2-7-5-7-9V6z"/>',
    "plus": '<circle cx="12" cy="12" r="9"/><path d="M12 8v8M8 12h8"/>',
    "list": '<path d="M8 6h13M8 12h13M8 18h13"/><circle cx="4" cy="6" r="1"/><circle cx="4" cy="12" r="1"/><circle cx="4" cy="18" r="1"/>',
    "users": '<circle cx="9" cy="8" r="3.4"/><path d="M3 20c0-3.3 2.7-6 6-6s6 2.7 6 6"/><path d="M16.5 5.2a3.4 3.4 0 010 6.6M21 20c0-2.4-1.4-4.5-3.5-5.4"/>',
    "chart": '<path d="M4 20h16"/><rect x="5" y="12" width="3.2" height="6" rx="1"/><rect x="10.4" y="7" width="3.2" height="11" rx="1"/><rect x="15.8" y="9.5" width="3.2" height="8.5" rx="1"/>',
    "pay": '<rect x="3" y="6" width="18" height="12" rx="2"/><path d="M3 10h18"/>',
    "user": '<circle cx="12" cy="8" r="4"/><path d="M5 21c0-3.9 3.1-7 7-7s7 3.1 7 7"/>',
    "key": '<circle cx="8" cy="8" r="4"/><path d="M11 11l9 9M17 17l2-2M19 15l2-2"/>',
    "award": '<circle cx="12" cy="9" r="5"/><path d="M9 13l-2 8 5-3 5 3-2-8"/>',
    "rules": '<circle cx="12" cy="12" r="9"/><path d="M8 12l3 3 5-6"/>',
    "doc": '<path d="M6 2h8l4 4v16H6z"/><path d="M14 2v4h4"/>',
    "alert": '<path d="M12 3l10 18H2z"/><path d="M12 10v5M12 18h.01"/>',
    "ruble": '<circle cx="12" cy="12" r="9"/><path d="M9 8h4a2.5 2.5 0 010 5H9v-5m0 5v4m-1 0h5m-6-2h4"/>',
    "flag": '<path d="M5 21V4h11l-2 4 2 4H5"/>',
    "wallet": '<rect x="3" y="6" width="18" height="13" rx="2"/><path d="M16 12h3"/>',
}


def ic(name):
    return Markup(f'<svg class="ic" viewBox="0 0 24 24" aria-hidden="true">'
                  f'{_ICONS.get(name, "")}</svg>')


def ictile(name):
    """Иконка в цветной плитке — для бокового меню."""
    return Markup(f'<span class="tile t-{name}"><svg class="ic" viewBox="0 0 24 24" '
                  f'aria-hidden="true">{_ICONS.get(name, "")}</svg></span>')


def asset_ver():
    """Версия статики по времени изменения style.css — чтобы браузер
    подхватывал новый CSS сам, без Ctrl+F5."""
    try:
        return int(os.path.getmtime(os.path.join(app.static_folder, "style.css")))
    except Exception:  # noqa: BLE001
        return 1


app.jinja_env.globals.update(FILE_KIND_RU=FILE_KIND_RU, FILE_KINDS=FILE_KINDS,
                             human_size=human_size, ic=ic, ictile=ictile,
                             asset_ver=asset_ver)
app.jinja_env.globals.update(STAGE_RU=STAGE_RU, SEV_RU=SEV_RU, KIND_RU=KIND_RU,
                             STAGE_ORDER=["draft", "internal_review", "legal_review",
                                          "at_counterparty", "in_progress", "completed",
                                          "warranty", "archived", "cancelled", "on_hold"])


@app.template_filter("money")
def money(v):
    if v is None:
        return "—"
    return f"{float(v):,.0f}".replace(",", " ") + " ₽"


@app.template_filter("dt")
def dt(v, fmt="%d.%m.%Y"):
    return v.strftime(fmt) if v else "—"


# ------------------------------------------------------------------
#  Аутентификация и права
# ------------------------------------------------------------------

def login_required(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if not session.get("user_id"):
            return redirect(url_for("login", next=request.path))
        return f(*a, **kw)
    return wrapper


def admin_required(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if not g.user or g.user["role_code"] != "admin":
            abort(403)
        return f(*a, **kw)
    return wrapper


@app.before_request
def load_user():
    g.user = None
    uid = session.get("user_id")
    if uid:
        g.user = db.query_one(
            "SELECT u.id, u.login, u.full_name, u.role_code, r.name AS role_name, "
            "u.must_change_password, d.name AS dept "
            "FROM users u JOIN roles r ON r.code=u.role_code "
            "LEFT JOIN departments d ON d.id=u.department_id WHERE u.id=%s", (uid,))
        # права роли
        g.perms = {r["cap_code"]: r["level"] for r in db.query(
            "SELECT cap_code, level FROM role_permissions WHERE role_code=%s",
            (g.user["role_code"],))}
        # заставляем сменить временный пароль до входа в остальные разделы
        if g.user and g.user["must_change_password"] and request.endpoint not in (
                "change_password", "logout", "static"):
            return redirect(url_for("change_password"))


def can(cap: str) -> bool:
    """Есть ли у текущего пользователя доступ к возможности (уровень не 'none')."""
    return g.get("perms", {}).get(cap, "none") != "none"


def perm_level(cap: str) -> str:
    return g.get("perms", {}).get(cap, "none")


app.jinja_env.globals.update(can=can, perm_level=perm_level)


def require_cap(cap):
    def deco(f):
        @wraps(f)
        def wrapper(*a, **kw):
            if not g.get("user"):
                return redirect(url_for("login", next=request.path))
            if not can(cap):
                abort(403)
            return f(*a, **kw)
        return wrapper
    return deco


def audit(action: str, entity: str, entity_id: int, before=None, after=None):
    """Запись в журнал действий (не удаляется и не правится триггерами)."""
    import json
    try:
        db.execute(
            "INSERT INTO audit_log(user_id, action, entity, entity_id, before, after, ip) "
            "VALUES(%s,%s,%s,%s,%s,%s,%s)",
            (g.user["id"] if g.get("user") else None, action, entity, entity_id,
             json.dumps(before, ensure_ascii=False, default=str) if before else None,
             json.dumps(after, ensure_ascii=False, default=str) if after else None,
             request.remote_addr))
    except Exception:  # noqa: BLE001
        pass


def next_number() -> str:
    """Следующий свободный номер вида М NN-MM-YYYY."""
    import datetime
    import re
    now = datetime.date.today()
    mx = 0
    for r in db.query("SELECT number_text FROM contracts"):
        m = re.match(r"М\s*(\d+)-", r["number_text"] or "")
        if m:
            mx = max(mx, int(m.group(1)))
    return f"М {mx + 1:02d}-{now.month:02d}-{now.year}"


@app.context_processor
def inject_today():
    import datetime
    return {"today": datetime.date.today()}


@app.context_processor
def inject_nav():
    """Счётчики для бокового меню."""
    if not g.get("user"):
        return {}
    counts = db.query_one("""
        SELECT
          (SELECT count(*) FROM contracts) AS contracts,
          (SELECT count(*) FROM contracts WHERE stage NOT IN ('archived','cancelled')) AS active,
          (SELECT count(*) FROM review_findings WHERE resolution IS NULL) AS findings,
          (SELECT count(*) FROM counterparties) AS counterparties,
          (SELECT count(*) FROM contracts WHERE warranty_months IS NOT NULL) AS warranty,
          (SELECT count(*) FROM users) AS users,
          (SELECT count(*) FROM stages s JOIN contracts c ON c.id=s.contract_id
             WHERE s.is_done=false AND s.planned_on IS NOT NULL AND s.planned_on < current_date
               AND c.stage NOT IN ('archived','cancelled')) AS stages_over,
          (SELECT count(*) FROM contracts
             WHERE valid_to IS NOT NULL AND valid_to BETWEEN current_date AND current_date + 30
               AND stage NOT IN ('archived','cancelled')) AS ending_soon,
          (SELECT count(*) FROM payments p JOIN contracts c ON c.id=p.contract_id
             WHERE p.paid_on IS NULL AND p.planned_on IS NOT NULL AND p.planned_on < current_date
               AND c.stage NOT IN ('archived','cancelled')) AS pay_over,
          (SELECT count(*) FROM tasks WHERE is_done=false AND assignee_id=%s) AS my_tasks,
          (SELECT count(*) FROM tasks WHERE is_done=false AND assignee_id=%s
             AND due_on IS NOT NULL AND due_on < current_date) AS my_tasks_over
    """, (g.user["id"], g.user["id"])) or {}
    show_pay = g.get("perms", {}).get("payments", "none") != "none"
    counts = dict(counts)
    counts["deadlines"] = (counts.get("stages_over", 0) + counts.get("ending_soon", 0)
                           + counts.get("my_tasks_over", 0)
                           + (counts.get("pay_over", 0) if show_pay else 0))
    return {"nav": counts}


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        row = db.query_one(
            "SELECT id, password_hash, is_active FROM users WHERE login=%s",
            ((request.form.get("login") or "").strip(),))
        if row and row["is_active"] and check_password_hash(
                row["password_hash"], request.form.get("password") or ""):
            session.clear()
            session["user_id"] = row["id"]
            db.execute("UPDATE users SET last_login_at=now() WHERE id=%s", (row["id"],))
            return redirect(request.args.get("next") or url_for("index"))
        flash("Неверный логин или пароль")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ------------------------------------------------------------------
#  Мой день (сводка)
# ------------------------------------------------------------------

@app.route("/")
@login_required
def index():
    stats = db.query_one("""
        SELECT
          count(*) AS total,
          count(*) FILTER (WHERE stage NOT IN ('archived','cancelled')) AS active,
          count(*) FILTER (WHERE stage='cancelled') AS cancelled,
          count(*) FILTER (WHERE stage='on_hold') AS on_hold,
          coalesce(sum(amount) FILTER (WHERE stage <> 'cancelled'),0) AS total_amount
        FROM contracts
    """) or {}
    sev = db.query_one("""
        SELECT
          count(*) FILTER (WHERE severity='critical') AS crit,
          count(*) FILTER (WHERE severity='warning') AS warn,
          count(*) FILTER (WHERE severity='info') AS info
        FROM review_findings WHERE resolution IS NULL
    """) or {}
    # договоры с критическими замечаниями — на стол
    critical = db.query("""
        SELECT c.id, c.number_text, cp.name AS counterparty,
               count(*) AS n, min(rf.title) AS sample
        FROM review_findings rf
        JOIN contracts c ON c.id=rf.contract_id
        LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
        WHERE rf.resolution IS NULL AND rf.severity='critical'
        GROUP BY c.id, c.number_text, cp.name
        ORDER BY c.number_text DESC
    """)
    warns = db.query("""
        SELECT c.id, c.number_text, cp.name AS counterparty,
               count(*) AS n, min(rf.title) AS sample
        FROM review_findings rf
        JOIN contracts c ON c.id=rf.contract_id
        LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
        WHERE rf.resolution IS NULL AND rf.severity='warning'
        GROUP BY c.id, c.number_text, cp.name
        ORDER BY c.number_text DESC LIMIT 12
    """)
    by_stage = {r["stage"]: r["n"] for r in db.query(
        "SELECT stage, count(*) AS n FROM contracts GROUP BY stage")}
    return render_template("dashboard.html", stats=stats, sev=sev,
                           critical=critical, warns=warns,
                           by_stage=by_stage, stage_order=STAGE_ORDER)


# ------------------------------------------------------------------
#  Доска стадий
# ------------------------------------------------------------------

@app.route("/board")
@login_required
def board():
    rows = db.query("""
        SELECT c.id, c.number_text, c.subject, c.amount, c.stage, c.stage_since,
               cp.name AS counterparty,
               (SELECT count(*) FROM review_findings rf
                 WHERE rf.contract_id=c.id AND rf.resolution IS NULL
                   AND rf.severity='critical') AS crit
        FROM contracts c LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
        ORDER BY c.stage_since DESC NULLS LAST, c.id DESC
    """)
    cols = {s: [] for s in STAGE_ORDER}
    for r in rows:
        cols.setdefault(r["stage"], []).append(r)
    return render_template("board.html", cols=cols, stage_order=STAGE_ORDER)


# ------------------------------------------------------------------
#  Реестр
# ------------------------------------------------------------------

@app.route("/registry")
@login_required
def registry():
    flt = request.args.get("f", "all")
    clauses = []
    params = []
    # быстрые вкладки
    if flt == "active":
        clauses.append("c.stage NOT IN ('archived','cancelled')")
    elif flt == "cancelled":
        clauses.append("c.stage='cancelled'")
    elif flt == "findings":
        clauses.append("EXISTS (SELECT 1 FROM review_findings rf "
                        "WHERE rf.contract_id=c.id AND rf.resolution IS NULL)")
    elif flt == "warranty":
        clauses.append("c.warranty_months IS NOT NULL")

    # параметрические фильтры
    q = (request.args.get("q") or "").strip()
    if q:
        clauses.append("(c.number_text ILIKE %s OR c.external_number ILIKE %s "
                       "OR cp.name ILIKE %s OR o.address ILIKE %s OR c.subject ILIKE %s)")
        params += [f"%{q}%"] * 5
    f_type = request.args.get("type") or ""
    if f_type:
        clauses.append("c.type_code = %s")
        params.append(f_type)
    f_cp = request.args.get("cp") or ""
    if f_cp.isdigit():
        clauses.append("c.counterparty_id = %s")
        params.append(int(f_cp))
    f_stage = request.args.get("stage") or ""
    if f_stage:
        clauses.append("c.stage = %s")
        params.append(f_stage)
    d_from = request.args.get("from") or ""
    if d_from:
        clauses.append("c.signed_on >= %s")
        params.append(d_from)
    d_to = request.args.get("to") or ""
    if d_to:
        clauses.append("c.signed_on <= %s")
        params.append(d_to)
    a_min = (request.args.get("amin") or "").replace(" ", "")
    if a_min.isdigit():
        clauses.append("c.amount >= %s")
        params.append(int(a_min))
    a_max = (request.args.get("amax") or "").replace(" ", "")
    if a_max.isdigit():
        clauses.append("c.amount <= %s")
        params.append(int(a_max))

    where = " AND ".join(clauses) if clauses else "1=1"
    rows = db.query(f"""
        SELECT c.id, c.number_text, c.subject, c.amount, c.advance_pct, c.stage,
               c.signed_on, c.warranty_months, cp.name AS counterparty,
               ct.name AS type_name, o.address AS object,
               (SELECT count(*) FROM review_findings rf
                 WHERE rf.contract_id=c.id AND rf.resolution IS NULL) AS findings
        FROM contracts c
        LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
        LEFT JOIN contract_types ct ON ct.code=c.type_code
        LEFT JOIN objects o ON o.id=c.object_id
        WHERE {where}
        ORDER BY c.number_text DESC LIMIT 500
    """, tuple(params))
    total_amount = sum(float(r["amount"]) for r in rows if r["amount"])
    types = db.query("SELECT code, name FROM contract_types ORDER BY name")
    cps = db.query("SELECT id, name FROM counterparties WHERE is_active ORDER BY name")
    args = request.args
    adv = any(args.get(k) for k in ("q", "type", "cp", "stage", "from", "to", "amin", "amax"))
    return render_template("registry.html", rows=rows, flt=flt, types=types, cps=cps,
                           args=args, adv=adv, total_amount=total_amount)


@app.route("/contract/<int:cid>")
@login_required
def contract(cid: int):
    c = db.query_one("""
        SELECT c.*, cp.name AS counterparty, cp.kind AS cp_kind,
               ct.name AS type_name, o.address AS object_address
        FROM contracts c
        LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
        LEFT JOIN contract_types ct ON ct.code=c.type_code
        LEFT JOIN objects o ON o.id=c.object_id
        WHERE c.id=%s
    """, (cid,))
    if not c:
        abort(404)
    findings = db.query(
        "SELECT id, severity, title, detail, clause, resolution FROM review_findings "
        "WHERE contract_id=%s ORDER BY CASE severity "
        "WHEN 'critical' THEN 0 WHEN 'warning' THEN 1 ELSE 2 END, id", (cid,))
    files = db.query("""
        SELECT cf.id, cf.kind, cf.version, cf.file_name, cf.size_bytes,
               cf.uploaded_at, u.full_name AS uploader
        FROM contract_files cf LEFT JOIN users u ON u.id=cf.uploaded_by
        WHERE cf.contract_id=%s
        ORDER BY cf.kind, cf.version DESC, cf.id DESC
    """, (cid,))
    work_stages = db.query("""
        SELECT id, ord, name, volume, planned_on, actual_on, amount, is_done
        FROM stages WHERE contract_id=%s ORDER BY ord, id
    """, (cid,))
    payments = []
    pay_sum = {}
    if perm_level("payments") != "none":
        payments = db.query("""
            SELECT id, kind, direction, planned_on, amount, paid_on, paid_amount, condition
            FROM payments WHERE contract_id=%s ORDER BY planned_on NULLS LAST, id
        """, (cid,))
        pay_sum = db.query_one("""
            SELECT coalesce(sum(amount),0) AS planned,
                   coalesce(sum(paid_amount) FILTER (WHERE paid_on IS NOT NULL),0) AS paid
            FROM payments WHERE contract_id=%s AND direction='incoming'
        """, (cid,)) or {}
    comments = db.query("""
        SELECT cm.id, cm.body, cm.created_at, u.full_name AS author
        FROM comments cm LEFT JOIN users u ON u.id=cm.author_id
        WHERE cm.contract_id=%s ORDER BY cm.created_at
    """, (cid,))
    tasks = db.query("""
        SELECT t.id, t.title, t.due_on, t.priority, t.is_done, t.done_at,
               (t.due_on - current_date) AS days_left,
               a.full_name AS assignee, d.full_name AS doneby
        FROM tasks t
        LEFT JOIN users a ON a.id=t.assignee_id
        LEFT JOIN users d ON d.id=t.done_by
        WHERE t.contract_id=%s
        ORDER BY t.is_done, t.due_on NULLS LAST, t.id
    """, (cid,))
    team = db.query("SELECT id, full_name FROM users WHERE is_active ORDER BY full_name")
    links = db.query("""
        SELECT l.id, l.link_type, l.comment, l.parent_id,
               (CASE WHEN l.parent_id=%s THEN l.child_id ELSE l.parent_id END) AS other_id,
               c2.number_text, c2.stage, cp.name AS counterparty
        FROM contract_links l
        JOIN contracts c2 ON c2.id = (CASE WHEN l.parent_id=%s THEN l.child_id ELSE l.parent_id END)
        LEFT JOIN counterparties cp ON cp.id=c2.counterparty_id
        WHERE l.parent_id=%s OR l.child_id=%s
        ORDER BY l.created_at DESC
    """, (cid, cid, cid, cid))
    link_targets = db.query("""
        SELECT c.id, c.number_text, cp.name AS counterparty
        FROM contracts c LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
        WHERE c.id<>%s ORDER BY c.number_text DESC LIMIT 400
    """, (cid,))
    approvals = db.query("""
        SELECT a.id, a.ord, a.status, a.comment, a.decided_at, a.approver_id,
               u.full_name AS approver
        FROM approvals a LEFT JOIN users u ON u.id=a.approver_id
        WHERE a.contract_id=%s ORDER BY a.ord, a.id
    """, (cid,))
    return render_template("contract.html", c=c, findings=findings, files=files,
                           payments=payments, pay_sum=pay_sum, comments=comments,
                           tasks=tasks, team=team, links=links, link_targets=link_targets,
                           approvals=approvals)


@app.route("/contract/<int:cid>/comment", methods=["POST"])
@login_required
def comment_add(cid):
    if not db.query_one("SELECT 1 FROM contracts WHERE id=%s", (cid,)):
        abort(404)
    body = (request.form.get("body") or "").strip()
    if body:
        db.execute("INSERT INTO comments(contract_id, author_id, body) VALUES(%s,%s,%s)",
                   (cid, g.user["id"], body[:4000]))
        audit("comment", "contract", cid, after={"body": body[:200]})
    return redirect(url_for("contract", cid=cid) + "#discuss")


# ------------------------------------------------------------------
#  Задачи и поручения
# ------------------------------------------------------------------

PRIORITY_RU = {"low": "низкий", "normal": "обычный", "high": "высокий"}
app.jinja_env.globals.update(PRIORITY_RU=PRIORITY_RU)


@app.route("/contract/<int:cid>/task", methods=["POST"])
@login_required
def task_add(cid):
    if not can("tasks"):
        abort(403)
    if not db.query_one("SELECT 1 FROM contracts WHERE id=%s", (cid,)):
        abort(404)
    title = (request.form.get("title") or "").strip()
    if not title:
        return redirect(url_for("contract", cid=cid) + "#tasks")
    assignee = request.form.get("assignee_id") or None
    due = request.form.get("due_on") or None
    priority = request.form.get("priority") or "normal"
    if priority not in PRIORITY_RU:
        priority = "normal"
    db.execute("""INSERT INTO tasks(contract_id, title, assignee_id, created_by, due_on, priority)
                  VALUES(%s,%s,%s,%s,%s,%s)""",
               (cid, title[:300], assignee, g.user["id"], due, priority))
    audit("task_add", "contract", cid, after={"title": title[:200], "due_on": due})
    return redirect(url_for("contract", cid=cid) + "#tasks")


@app.route("/task/<int:tid>/done", methods=["POST"])
@login_required
def task_done(tid):
    if not can("tasks"):
        abort(403)
    t = db.query_one("SELECT id, contract_id, is_done FROM tasks WHERE id=%s", (tid,))
    if not t:
        abort(404)
    new_done = not t["is_done"]
    db.execute("""UPDATE tasks SET is_done=%s,
                    done_at = CASE WHEN %s THEN now() ELSE NULL END,
                    done_by = CASE WHEN %s THEN %s ELSE NULL END
                  WHERE id=%s""",
               (new_done, new_done, new_done, g.user["id"], tid))
    audit("task_done" if new_done else "task_reopen", "contract",
          t["contract_id"], after={"task_id": tid})
    back = request.form.get("back") or (url_for("contract", cid=t["contract_id"]) + "#tasks")
    return redirect(back)


@app.route("/task/<int:tid>/delete", methods=["POST"])
@login_required
def task_delete(tid):
    if not can("tasks"):
        abort(403)
    t = db.query_one("SELECT id, contract_id FROM tasks WHERE id=%s", (tid,))
    if not t:
        abort(404)
    db.execute("DELETE FROM tasks WHERE id=%s", (tid,))
    audit("task_delete", "contract", t["contract_id"], after={"task_id": tid})
    return redirect(url_for("contract", cid=t["contract_id"]) + "#tasks")


@app.route("/tasks")
@login_required
def tasks_page():
    if not can("tasks"):
        abort(403)
    flt = request.args.get("f", "my")
    uid = g.user["id"]
    own_only = perm_level("tasks") == "own"
    where = "t.is_done = false"
    params = []
    if flt == "my" or own_only:
        where += " AND t.assignee_id = %s"
        params.append(uid)
    elif flt == "overdue":
        where += " AND t.due_on IS NOT NULL AND t.due_on < current_date"
    elif flt == "done":
        where = "t.is_done = true"
        if own_only:
            where += " AND t.assignee_id = %s"
            params.append(uid)
    rows = db.query(f"""
        SELECT t.id, t.title, t.due_on, t.priority, t.is_done, t.contract_id,
               (t.due_on - current_date) AS days_left,
               a.full_name AS assignee, c.number_text, cp.name AS counterparty
        FROM tasks t
        LEFT JOIN users a ON a.id=t.assignee_id
        LEFT JOIN contracts c ON c.id=t.contract_id
        LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
        WHERE {where}
        ORDER BY t.is_done, t.due_on NULLS LAST, t.priority DESC, t.id
    """, tuple(params))
    return render_template("tasks.html", rows=rows, flt=flt, own_only=own_only)


# ------------------------------------------------------------------
#  Контрагенты
# ------------------------------------------------------------------

PAY_KIND_RU = {"advance": "аванс", "stage": "этап", "final": "окончательный"}
DIR_RU = {"incoming": "нам платят", "outgoing": "мы платим"}
app.jinja_env.globals.update(PAY_KIND_RU=PAY_KIND_RU, DIR_RU=DIR_RU)


@app.route("/payments")
@login_required
def payments_page():
    if perm_level("payments") == "none":
        abort(403)
    flt = request.args.get("f", "open")
    where = "p.paid_on IS NULL"
    if flt == "overdue":
        where = "p.paid_on IS NULL AND p.planned_on < current_date"
    elif flt == "incoming":
        where = "p.paid_on IS NULL AND p.direction='incoming'"
    elif flt == "outgoing":
        where = "p.paid_on IS NULL AND p.direction='outgoing'"
    elif flt == "paid":
        where = "p.paid_on IS NOT NULL"
    rows = db.query(f"""
        SELECT p.id, p.kind, p.direction, p.planned_on, p.amount, p.paid_on,
               p.condition, c.id AS cid, c.number_text, cp.name AS counterparty
        FROM payments p JOIN contracts c ON c.id=p.contract_id
        LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
        WHERE {where}
        ORDER BY p.planned_on NULLS LAST, p.id
    """)
    totals = db.query_one("""
        SELECT
          coalesce(sum(amount) FILTER (WHERE paid_on IS NULL AND direction='incoming'),0) AS to_get,
          coalesce(sum(amount) FILTER (WHERE paid_on IS NULL AND direction='outgoing'),0) AS to_pay,
          count(*) FILTER (WHERE paid_on IS NULL AND planned_on < current_date) AS overdue
        FROM payments
    """) or {}
    return render_template("payments.html", rows=rows, flt=flt, totals=totals)


@app.route("/contract/<int:cid>/payment", methods=["POST"])
@login_required
def payment_add(cid):
    if perm_level("payments") != "yes":
        abort(403)
    kind = request.form.get("kind", "final")
    direction = request.form.get("direction", "incoming")
    planned = request.form.get("planned_on") or None
    amount = request.form.get("amount") or None
    condition = (request.form.get("condition") or "").strip() or None
    if kind not in PAY_KIND_RU:
        kind = "final"
    if direction not in DIR_RU:
        direction = "incoming"
    if not amount:
        flash("Укажите сумму платежа")
        return redirect(url_for("contract", cid=cid))
    pid = db.query_one(
        "INSERT INTO payments(contract_id, kind, direction, planned_on, amount, condition) "
        "VALUES(%s,%s,%s,%s,%s,%s) RETURNING id",
        (cid, kind, direction, planned, amount, condition))["id"]
    audit("payment_add", "contract", cid, after={"amount": amount, "kind": kind})
    flash("Платёж добавлен в график")
    return redirect(url_for("contract", cid=cid))


@app.route("/payment/<int:pid>/paid", methods=["POST"])
@login_required
def payment_paid(pid):
    if perm_level("payments") != "yes":
        abort(403)
    p = db.query_one("SELECT contract_id, amount FROM payments WHERE id=%s", (pid,))
    if not p:
        abort(404)
    if request.form.get("undo") == "1":
        db.execute("UPDATE payments SET paid_on=NULL, paid_amount=NULL WHERE id=%s", (pid,))
    else:
        paid_on = request.form.get("paid_on") or None
        paid_amount = request.form.get("paid_amount") or p["amount"]
        db.execute("UPDATE payments SET paid_on=coalesce(%s, current_date), paid_amount=%s WHERE id=%s",
                   (paid_on, paid_amount, pid))
    audit("payment_paid", "contract", p["contract_id"], after={"payment": pid})
    return redirect(url_for("contract", cid=p["contract_id"]))


@app.route("/contract/<int:cid>/payment/schedule", methods=["POST"])
@login_required
def payment_schedule(cid):
    """Быстрый график: аванс (если есть) + остаток."""
    if perm_level("payments") != "yes":
        abort(403)
    c = db.query_one("SELECT amount, advance_amount, advance_pct FROM contracts WHERE id=%s", (cid,))
    if not c or not c["amount"]:
        flash("У договора не указана сумма — график не построить")
        return redirect(url_for("contract", cid=cid))
    if db.query_one("SELECT 1 FROM payments WHERE contract_id=%s LIMIT 1", (cid,)):
        flash("График уже есть — добавляйте платежи вручную")
        return redirect(url_for("contract", cid=cid))
    from decimal import Decimal
    total = Decimal(c["amount"])
    adv = Decimal(c["advance_amount"]) if c["advance_amount"] else Decimal(0)
    if adv > 0:
        db.execute("INSERT INTO payments(contract_id, kind, direction, amount, condition) "
                   "VALUES(%s,'advance','incoming',%s,%s)",
                   (cid, adv, "аванс по договору"))
    rest = total - adv
    if rest > 0:
        db.execute("INSERT INTO payments(contract_id, kind, direction, amount, condition) "
                   "VALUES(%s,'final','incoming',%s,%s)",
                   (cid, rest, "окончательный расчёт по акту"))
    audit("payment_schedule", "contract", cid)
    flash("Простой график создан: аванс и остаток")
    return redirect(url_for("contract", cid=cid))


@app.route("/counterparty/<int:pid>")
@login_required
def counterparty(pid):
    cp = db.query_one("SELECT * FROM counterparties WHERE id=%s", (pid,))
    if not cp:
        abort(404)
    m = db.query_one("""
        SELECT count(*) AS total,
               count(*) FILTER (WHERE stage='cancelled') AS cancelled,
               count(*) FILTER (WHERE stage='completed') AS completed,
               count(*) FILTER (WHERE stage NOT IN ('cancelled','archived','completed')) AS active,
               count(*) FILTER (WHERE warranty_months IS NOT NULL) AS warranty,
               coalesce(sum(amount) FILTER (WHERE stage<>'cancelled'),0) AS total_amount,
               (SELECT count(*) FROM review_findings rf JOIN contracts c2 ON c2.id=rf.contract_id
                 WHERE c2.counterparty_id=%s AND rf.resolution IS NULL) AS findings
        FROM contracts WHERE counterparty_id=%s
    """, (pid, pid)) or {}
    contracts = db.query("""
        SELECT c.id, c.number_text, c.subject, c.amount, c.stage, c.signed_on,
               o.address AS object,
               (SELECT count(*) FROM review_findings rf WHERE rf.contract_id=c.id AND rf.resolution IS NULL) AS findings
        FROM contracts c LEFT JOIN objects o ON o.id=c.object_id
        WHERE c.counterparty_id=%s ORDER BY c.number_text DESC
    """, (pid,))
    total = m.get("total") or 0
    cancelled = m.get("cancelled") or 0
    reliability = round((total - cancelled) / total * 100) if total else None
    return render_template("counterparty.html", cp=cp, m=m, contracts=contracts,
                           reliability=reliability)


@app.route("/counterparties")
@login_required
def counterparties():
    rows = db.query("""
        SELECT cp.id, cp.name, cp.kind, count(c.id) AS n,
               coalesce(sum(c.amount),0) AS total,
               count(*) FILTER (WHERE c.stage='cancelled') AS cancelled
        FROM counterparties cp LEFT JOIN contracts c ON c.counterparty_id=cp.id
        GROUP BY cp.id, cp.name, cp.kind
        ORDER BY count(c.id) DESC, cp.name
    """)
    return render_template("counterparties.html", rows=rows)


# ------------------------------------------------------------------
#  Гарантии
# ------------------------------------------------------------------

@app.route("/warranty")
@login_required
def warranty():
    rows = db.query("""
        SELECT c.id, c.number_text, c.warranty_months, c.warranty_until,
               c.commissioning_date, c.type_code, cp.name AS counterparty,
               o.address AS object
        FROM contracts c
        LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
        LEFT JOIN objects o ON o.id=c.object_id
        WHERE c.warranty_months IS NOT NULL
        ORDER BY c.warranty_months DESC, c.number_text
    """)
    return render_template("warranty.html", rows=rows)


# ------------------------------------------------------------------
#  Пользователи и права
# ------------------------------------------------------------------

@app.route("/users")
@require_cap("manage_users")
def users():
    rows = db.query("""
        SELECT u.id, u.login, u.full_name, u.role_code, r.name AS role_name,
               d.name AS dept, u.is_active, u.last_login_at
        FROM users u JOIN roles r ON r.code=u.role_code
        LEFT JOIN departments d ON d.id=u.department_id
        ORDER BY u.id
    """)
    return render_template("users.html", rows=rows)


LEVELS = [("none", "нет"), ("own", "свои"), ("dept", "свой отдел"),
          ("view", "просмотр"), ("all", "все"), ("yes", "да")]


def role_order():
    return ("CASE code WHEN 'manager' THEN 0 WHEN 'lawyer' THEN 1 "
            "WHEN 'head' THEN 2 WHEN 'accountant' THEN 3 WHEN 'admin' THEN 5 ELSE 4 END")


@app.route("/permissions", methods=["GET", "POST"])
@require_cap("manage_permissions")
def permissions():
    roles = db.query(f"SELECT code, name, is_builtin FROM roles ORDER BY {role_order()}")
    caps = db.query("SELECT code, name FROM capabilities ORDER BY ord")
    if request.method == "POST":
        for cap in caps:
            for role in roles:
                if role["code"] == "admin":
                    continue  # админ всегда может всё
                key = f"{role['code']}__{cap['code']}"
                if key not in request.form:
                    continue  # не трогаем поля, которых нет в форме
                level = request.form.get(key, "none")
                if level not in dict(LEVELS):
                    level = "none"
                db.execute(
                    "INSERT INTO role_permissions(role_code, cap_code, level) VALUES(%s,%s,%s) "
                    "ON CONFLICT (role_code, cap_code) DO UPDATE SET level=EXCLUDED.level",
                    (role["code"], cap["code"], level))
        audit("permissions", "role_permissions", 0)
        flash("Права сохранены")
        return redirect(url_for("permissions"))
    matrix = {}
    for r in db.query("SELECT role_code, cap_code, level FROM role_permissions"):
        matrix[(r["role_code"], r["cap_code"])] = r["level"]
    return render_template("permissions.html", roles=roles, caps=caps,
                           matrix=matrix, levels=LEVELS)


@app.route("/roles")
@require_cap("manage_permissions")
def roles_page():
    rows = db.query(f"""
        SELECT r.code, r.name, r.description, r.is_builtin,
               (SELECT count(*) FROM users u WHERE u.role_code=r.code) AS users
        FROM roles r ORDER BY {role_order()}""")
    return render_template("roles.html", rows=rows)


@app.route("/roles/new", methods=["POST"])
@require_cap("manage_permissions")
def role_new():
    import re
    name = (request.form.get("name") or "").strip()
    desc = (request.form.get("description") or "").strip()
    code = (request.form.get("code") or "").strip().lower()
    if not code:
        code = re.sub(r"[^a-z0-9_]", "", (name.lower().replace(" ", "_")))[:20] or None
    if not name or not code:
        flash("Укажите название роли")
    elif db.query_one("SELECT 1 FROM roles WHERE code=%s", (code,)):
        flash("Роль с таким кодом уже есть")
    else:
        db.execute("INSERT INTO roles(code, name, description, is_builtin) VALUES(%s,%s,%s,false)",
                   (code, name, desc))
        # новой роли — всё запрещено по умолчанию
        for cap in db.query("SELECT code FROM capabilities"):
            db.execute("INSERT INTO role_permissions(role_code, cap_code, level) VALUES(%s,%s,'none') "
                       "ON CONFLICT DO NOTHING", (code, cap["code"]))
        flash(f"Роль «{name}» создана — настройте её права в матрице")
    return redirect(url_for("roles_page"))


@app.route("/roles/<code>/edit", methods=["POST"])
@require_cap("manage_permissions")
def role_edit(code):
    name = (request.form.get("name") or "").strip()
    desc = (request.form.get("description") or "").strip()
    if name:
        db.execute("UPDATE roles SET name=%s, description=%s WHERE code=%s", (name, desc, code))
        flash("Роль обновлена")
    return redirect(url_for("roles_page"))


@app.route("/roles/<code>/delete", methods=["POST"])
@require_cap("manage_permissions")
def role_delete(code):
    r = db.query_one("SELECT is_builtin FROM roles WHERE code=%s", (code,))
    n = db.query_one("SELECT count(*) AS n FROM users WHERE role_code=%s", (code,))["n"]
    if not r or r["is_builtin"]:
        flash("Встроенную роль удалить нельзя")
    elif n:
        flash(f"Нельзя удалить: роль назначена {n} пользователям")
    else:
        db.execute("DELETE FROM roles WHERE code=%s", (code,))
        flash("Роль удалена")
    return redirect(url_for("roles_page"))


# ------------------------------------------------------------------
#  Служебное
# ------------------------------------------------------------------

@app.route("/account/password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        old = request.form.get("old") or ""
        new = request.form.get("new") or ""
        new2 = request.form.get("new2") or ""
        row = db.query_one("SELECT password_hash FROM users WHERE id=%s", (g.user["id"],))
        forced = g.user["must_change_password"]
        if not forced and not check_password_hash(row["password_hash"], old):
            flash("Текущий пароль неверный")
        elif len(new) < 8:
            flash("Новый пароль слишком короткий — минимум 8 символов")
        elif new != new2:
            flash("Пароли не совпадают")
        else:
            db.execute("UPDATE users SET password_hash=%s, must_change_password=false WHERE id=%s",
                       (generate_password_hash(new), g.user["id"]))
            flash("Пароль изменён")
            return redirect(url_for("index"))
    return render_template("password.html", forced=g.user["must_change_password"])


@app.route("/users/new", methods=["GET", "POST"])
@require_cap("manage_users")
def user_new():
    depts = db.query("SELECT id, name FROM departments ORDER BY name")
    roles = db.query("SELECT code, name FROM roles ORDER BY name")
    if request.method == "POST":
        login_name = (request.form.get("login") or "").strip()
        full_name = (request.form.get("full_name") or "").strip()
        role = request.form.get("role") or "manager"
        dept = request.form.get("department") or None
        if not login_name or not full_name:
            flash("Заполните имя и логин")
        elif db.query_one("SELECT 1 FROM users WHERE login=%s", (login_name,)):
            flash("Такой логин уже есть")
        else:
            import secrets, string
            pw = "".join(secrets.choice(string.ascii_letters + string.digits) for _ in range(10))
            db.execute(
                "INSERT INTO users(login, full_name, role_code, department_id, password_hash, "
                "must_change_password, is_active) VALUES(%s,%s,%s,%s,%s,true,true)",
                (login_name, full_name, role, dept or None, generate_password_hash(pw)))
            return render_template("user_created.html", login=login_name, password=pw)
    return render_template("user_new.html", depts=depts, roles=roles)


@app.route("/users/<int:uid>/toggle", methods=["POST"])
@require_cap("manage_users")
def user_toggle(uid):
    if uid != g.user["id"]:
        db.execute("UPDATE users SET is_active = NOT is_active WHERE id=%s", (uid,))
        audit("user_toggle", "user", uid)
    return redirect(url_for("users"))


@app.route("/users/<int:uid>/edit", methods=["GET", "POST"])
@require_cap("manage_users")
def user_edit(uid):
    u = db.query_one("SELECT * FROM users WHERE id=%s", (uid,))
    if not u:
        abort(404)
    depts = db.query("SELECT id, name FROM departments ORDER BY name")
    roles = db.query("SELECT code, name FROM roles ORDER BY name")
    if request.method == "POST":
        action = request.form.get("action")
        if action == "reset_password":
            import secrets as _s, string
            pw = "".join(_s.choice(string.ascii_letters + string.digits) for _ in range(10))
            db.execute("UPDATE users SET password_hash=%s, must_change_password=true WHERE id=%s",
                       (generate_password_hash(pw), uid))
            audit("user_reset_pw", "user", uid)
            return render_template("user_created.html", login=u["login"], password=pw)
        full_name = (request.form.get("full_name") or "").strip()
        login_name = (request.form.get("login") or "").strip()
        role = request.form.get("role") or u["role_code"]
        dept = request.form.get("department") or None
        is_active = request.form.get("is_active") == "on"
        if not full_name or not login_name:
            flash("Имя и логин обязательны")
        elif db.query_one("SELECT 1 FROM users WHERE login=%s AND id<>%s", (login_name, uid)):
            flash("Такой логин уже занят")
        else:
            # нельзя снять с себя админ-роль или отключить себя — чтобы не потерять доступ
            if uid == g.user["id"] and (role != "admin" or not is_active):
                flash("Нельзя снять права администратора или отключить самого себя")
            else:
                db.execute(
                    "UPDATE users SET full_name=%s, login=%s, role_code=%s, department_id=%s, "
                    "is_active=%s WHERE id=%s",
                    (full_name, login_name, role, dept or None, is_active, uid))
                audit("user_edit", "user", uid, after={"login": login_name, "role": role})
                flash("Пользователь сохранён")
                return redirect(url_for("users"))
    return render_template("user_edit.html", u=u, depts=depts, roles=roles)


@app.route("/contract/new", methods=["GET", "POST"])
@login_required
def contract_new():
    if not can("create_contract"):
        abort(403)
    types = db.query("SELECT code, name, default_advance_pct, warranty_months FROM contract_types ORDER BY name")
    cps = db.query("SELECT id, name FROM counterparties ORDER BY name")
    if request.method == "POST":
        type_code = request.form.get("type_code")
        cp_id = request.form.get("counterparty_id") or None
        cp_new = (request.form.get("counterparty_new") or "").strip()
        cp_kind = request.form.get("cp_kind") or "commercial"
        address = (request.form.get("address") or "").strip()
        subject = (request.form.get("subject") or "").strip()
        amount = request.form.get("amount") or None
        advance_pct = request.form.get("advance_pct") or None
        warranty = request.form.get("warranty_months") or None
        has_penalty = True if request.form.get("has_penalty") == "on" else False

        if not type_code or not subject:
            flash("Укажите тип и предмет договора")
        else:
            if cp_new and not cp_id:
                cp_id = db.query_one(
                    "INSERT INTO counterparties(name, kind) VALUES(%s,%s) RETURNING id",
                    (cp_new, cp_kind))["id"]
            obj_id = None
            if address:
                obj = db.query_one("SELECT id FROM objects WHERE address=%s", (address,))
                obj_id = obj["id"] if obj else db.query_one(
                    "INSERT INTO objects(address) VALUES(%s) RETURNING id", (address,))["id"]
            advance_amount = None
            if amount and advance_pct:
                advance_amount = round(float(amount) * float(advance_pct) / 100, 2)
            num = next_number()
            cid = db.query_one("""
                INSERT INTO contracts(number_text, type_code, counterparty_id, object_id,
                    subject, amount, advance_pct, advance_amount, warranty_months,
                    has_penalty, stage, created_by)
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'draft',%s) RETURNING id""",
                (num, type_code, cp_id, obj_id, subject, amount or None, advance_pct or None,
                 advance_amount, warranty or None, has_penalty, g.user["id"]))["id"]
            n = review_engine.run(cid)
            audit("create", "contract", cid, after={"number": num})
            flash(f"Договор {num} создан. Проверка нашла замечаний: {n}")
            return redirect(url_for("contract", cid=cid))
    return render_template("contract_new.html", types=types, cps=cps, next_num=next_number())


@app.route("/contract/<int:cid>/review", methods=["POST"])
@login_required
def contract_review(cid):
    n = review_engine.run(cid)
    audit("review", "contract", cid, after={"added": n})
    flash(f"Проверка добавила замечаний: {n}" if n else "Проверка завершена, новых замечаний нет")
    return redirect(url_for("contract", cid=cid))


STAGES = ["draft", "internal_review", "legal_review", "at_counterparty",
          "in_progress", "completed", "warranty", "archived", "cancelled", "on_hold"]


@app.route("/rules-catalog", methods=["GET", "POST"])
@require_cap("manage_permissions")
def rules_catalog():
    if request.method == "POST":
        all_rules = db.query("SELECT code FROM review_rules")
        for r in all_rules:
            on = request.form.get("rule_" + r["code"]) == "on"
            db.execute("UPDATE review_rules SET enabled=%s WHERE code=%s", (on, r["code"]))
        audit("rules", "review_rules", 0)
        flash("Набор правил сохранён")
        return redirect(url_for("rules_catalog"))
    rules = db.query("SELECT code, name, description, severity, category, needs_text, enabled "
                     "FROM review_rules ORDER BY ord")
    groups = {}
    for r in rules:
        groups.setdefault(r["category"], []).append(r)
    return render_template("rules_catalog.html", groups=groups)


@app.route("/contract/<int:cid>/ai_review", methods=["POST"])
@login_required
def contract_ai_review(cid):
    if not can("review_decide") and g.user["role_code"] != "admin":
        abort(403)
    # ИИ-проверка требует: текст договора в системе + ключ API + согласие.
    # Пока не настроена — честно сообщаем и не отправляем ничего наружу.
    if not os.getenv("AI_API_KEY"):
        flash("Умная проверка ИИ ещё не подключена: нужен ключ API и распознанный "
              "текст договора. Для секретных договоров она будет отключена.")
        return redirect(url_for("contract", cid=cid))
    flash("Умная проверка ИИ пока в разработке")
    return redirect(url_for("contract", cid=cid))


@app.route("/contract/<int:cid>/stage", methods=["POST"])
@login_required
def contract_stage(cid):
    if not can("change_stage"):
        abort(403)
    new_stage = request.form.get("stage")
    if new_stage not in STAGES:
        abort(400)
    old = db.query_one("SELECT stage FROM contracts WHERE id=%s", (cid,))
    if old and old["stage"] == new_stage:
        flash("Стадия не изменилась")
        return redirect(url_for("contract", cid=cid))
    db.execute("UPDATE contracts SET stage=%s, stage_since=now(), updated_at=now() WHERE id=%s",
               (new_stage, cid))
    audit("stage", "contract", cid, before=old,
          after={"stage": new_stage, "from": STAGE_RU.get(old["stage"] if old else "", "")})
    flash(f"Стадия изменена: {STAGE_RU.get(new_stage, new_stage)}")
    return redirect(url_for("contract", cid=cid))


@app.route("/contract/<int:cid>/move", methods=["POST"])
@login_required
def contract_move(cid):
    """Перетаскивание карточки на доске: смена стадии через AJAX + журнал."""
    if not can("change_stage"):
        return {"ok": False, "error": "Нет права менять стадию"}, 403
    new_stage = request.form.get("stage")
    if new_stage not in STAGES:
        return {"ok": False, "error": "Неизвестная стадия"}, 400
    old = db.query_one("SELECT stage FROM contracts WHERE id=%s", (cid,))
    if not old:
        return {"ok": False, "error": "Договор не найден"}, 404
    if old["stage"] == new_stage:
        return {"ok": True, "stage": new_stage, "unchanged": True}
    db.execute("UPDATE contracts SET stage=%s, stage_since=now(), updated_at=now() WHERE id=%s",
               (new_stage, cid))
    audit("stage_move", "contract", cid, before={"stage": old["stage"]},
          after={"stage": new_stage, "from_ru": STAGE_RU.get(old["stage"], ""),
                 "to_ru": STAGE_RU.get(new_stage, "")})
    import datetime
    return {"ok": True, "stage": new_stage,
            "since": datetime.date.today().strftime("%d.%m.%y"),
            "by": g.user["full_name"]}


@app.route("/contract/<int:cid>/edit", methods=["GET", "POST"])
@login_required
def contract_edit(cid):
    if not can("edit_contract"):
        abort(403)
    c = db.query_one("SELECT * FROM contracts WHERE id=%s", (cid,))
    if not c:
        abort(404)
    types = db.query("SELECT code, name FROM contract_types ORDER BY name")
    if request.method == "POST":
        f = request.form
        amount = f.get("amount") or None
        adv = f.get("advance_pct") or None
        advance_amount = round(float(amount) * float(adv) / 100, 2) if (amount and adv) else None
        db.execute("""
            UPDATE contracts SET subject=%s, amount=%s, advance_pct=%s, advance_amount=%s,
                warranty_months=%s, commissioning_date=%s, signed_on=%s,
                has_penalty=%s, updated_at=now() WHERE id=%s""",
            (f.get("subject"), amount, adv, advance_amount,
             f.get("warranty_months") or None, f.get("commissioning_date") or None,
             f.get("signed_on") or None, f.get("has_penalty") == "on", cid))
        audit("edit", "contract", cid)
        flash("Договор сохранён")
        return redirect(url_for("contract", cid=cid))
    return render_template("contract_edit.html", c=c, types=types)


@app.route("/contract/<int:cid>/execution", methods=["POST"])
@login_required
def contract_execution(cid):
    if not can("execution"):
        abort(403)
    comm = request.form.get("commissioning_date") or None
    comp = request.form.get("completed_on") or None
    db.execute(
        "UPDATE contracts SET commissioning_date=%s, completed_on=%s, updated_at=now() WHERE id=%s",
        (comm, comp, cid))
    audit("execution", "contract", cid,
          after={"commissioning": comm, "completed": comp})
    flash("Даты исполнения сохранены")
    return redirect(url_for("contract", cid=cid))


@app.route("/contract/<int:cid>/work-stage", methods=["POST"])
@login_required
def work_stage_add(cid):
    if not can("execution") and not can("edit_contract"):
        abort(403)
    name = (request.form.get("name") or "").strip()
    volume = (request.form.get("volume") or "").strip() or None
    planned = request.form.get("planned_on") or None
    amount = request.form.get("amount") or None
    if not name:
        flash("Укажите название этапа")
        return redirect(url_for("contract", cid=cid))
    ordn = (db.query_one("SELECT coalesce(max(ord),0)+1 AS o FROM stages WHERE contract_id=%s",
                         (cid,)) or {"o": 1})["o"]
    db.execute("INSERT INTO stages(contract_id, ord, name, volume, planned_on, amount) "
               "VALUES(%s,%s,%s,%s,%s,%s)", (cid, ordn, name, volume, planned, amount))
    audit("stage_add", "contract", cid, after={"name": name})
    flash("Этап работ добавлен")
    return redirect(url_for("contract", cid=cid))


@app.route("/work-stage/<int:sid>/done", methods=["POST"])
@login_required
def work_stage_done(sid):
    if not can("execution") and not can("edit_contract"):
        abort(403)
    s = db.query_one("SELECT contract_id, is_done FROM stages WHERE id=%s", (sid,))
    if not s:
        abort(404)
    if s["is_done"]:
        db.execute("UPDATE stages SET is_done=false, actual_on=NULL WHERE id=%s", (sid,))
    else:
        actual = request.form.get("actual_on") or None
        db.execute("UPDATE stages SET is_done=true, actual_on=coalesce(%s, current_date) WHERE id=%s",
                   (actual, sid))
    audit("stage_done", "contract", s["contract_id"], after={"stage": sid})
    return redirect(url_for("contract", cid=s["contract_id"]))


@app.route("/work-stage/<int:sid>/delete", methods=["POST"])
@login_required
def work_stage_delete(sid):
    if not can("execution") and not can("edit_contract"):
        abort(403)
    s = db.query_one("SELECT contract_id FROM stages WHERE id=%s", (sid,))
    if s:
        db.execute("DELETE FROM stages WHERE id=%s", (sid,))
        audit("stage_del", "contract", s["contract_id"])
        return redirect(url_for("contract", cid=s["contract_id"]))
    abort(404)


@app.route("/finding/<int:fid>/resolve", methods=["POST"])
@login_required
def finding_resolve(fid):
    if not can("review_decide"):
        abort(403)
    decision = request.form.get("decision")  # accepted | rejected | reopen
    f = db.query_one("SELECT contract_id FROM review_findings WHERE id=%s", (fid,))
    if not f:
        abort(404)
    if decision == "reopen":
        db.execute("UPDATE review_findings SET resolution=NULL, resolved_by=NULL, "
                   "resolved_at=NULL WHERE id=%s", (fid,))
    elif decision in ("accepted", "rejected"):
        db.execute("UPDATE review_findings SET resolution=%s, resolved_by=%s, resolved_at=now() "
                   "WHERE id=%s", (decision, g.user["id"], fid))
    audit("finding", "review_finding", fid, after={"decision": decision})
    return redirect(url_for("contract", cid=f["contract_id"]))


@app.route("/search")
@login_required
def search():
    q = (request.args.get("q") or "").strip()
    rows = []
    if q:
        like = f"%{q}%"
        rows = db.query("""
            SELECT c.id, c.number_text, c.subject, c.amount, c.stage, c.signed_on,
                   cp.name AS counterparty, o.address AS object,
                   (SELECT count(*) FROM review_findings rf
                     WHERE rf.contract_id=c.id AND rf.resolution IS NULL) AS findings,
                   EXISTS (SELECT 1 FROM contract_files cf
                     WHERE cf.contract_id=c.id
                       AND cf.text_search @@ plainto_tsquery('russian', %s)) AS in_text
            FROM contracts c
            LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
            LEFT JOIN objects o ON o.id=c.object_id
            WHERE c.number_text ILIKE %s OR c.subject ILIKE %s
               OR cp.name ILIKE %s OR o.address ILIKE %s
               OR c.external_number ILIKE %s
               OR EXISTS (SELECT 1 FROM contract_files cf
                     WHERE cf.contract_id=c.id
                       AND cf.text_search @@ plainto_tsquery('russian', %s))
            ORDER BY c.number_text DESC LIMIT 100
        """, (q, like, like, like, like, like, q))
    return render_template("search.html", q=q, rows=rows)


@app.route("/contract/<int:cid>/upload", methods=["POST"])
@login_required
def contract_upload(cid):
    if not can("manage_files"):
        abort(403)
    c = db.query_one("SELECT id FROM contracts WHERE id=%s", (cid,))
    if not c:
        abort(404)
    f = request.files.get("file")
    kind = request.form.get("kind", "other")
    if kind not in FILE_KIND_RU:
        kind = "other"
    if not f or not f.filename:
        flash("Файл не выбран")
        return redirect(url_for("contract", cid=cid))
    data = f.read()
    if not data:
        flash("Пустой файл")
        return redirect(url_for("contract", cid=cid))
    sha = hashlib.sha256(data).hexdigest()
    # уже загружали ровно этот файл?
    dup = db.query_one("SELECT id FROM contract_files WHERE contract_id=%s AND sha256=%s",
                       (cid, sha))
    if dup:
        flash("Такой файл уже загружен (совпадает содержимое)")
        return redirect(url_for("contract", cid=cid))
    ext = os.path.splitext(f.filename)[1][:12]
    disk = uuid.uuid4().hex + ext
    folder = os.path.join(FILES_DIR, str(cid))
    os.makedirs(folder, exist_ok=True)
    with open(os.path.join(folder, disk), "wb") as out:
        out.write(data)
    ver = (db.query_one(
        "SELECT coalesce(max(version),0)+1 AS v FROM contract_files WHERE contract_id=%s AND kind=%s",
        (cid, kind)) or {"v": 1})["v"]
    full = os.path.join(folder, disk)
    text = text_extract.extract(full, f.filename)
    has_text = bool(text)
    fid = db.query_one(
        "INSERT INTO contract_files(contract_id, version, kind, file_name, file_path, "
        "size_bytes, sha256, uploaded_by, has_text_layer, extracted_text) "
        "VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (cid, ver, kind, f.filename, full, len(data), sha,
         g.user["id"], has_text, text or None))["id"]
    audit("upload", "contract", cid,
          after={"file": f.filename, "kind": kind, "version": ver})
    msg = f"Файл загружен: {f.filename} (версия {ver})"
    if has_text and kind == "contract":
        n = review_engine.run(cid)
        if n:
            msg += f". Проверка по тексту нашла замечаний: {n}"
    elif ext.lower() in (".pdf", ".docx") and not has_text:
        msg += ". Текст не распознан (скан без текстового слоя)"
    flash(msg)
    return redirect(url_for("contract", cid=cid))


@app.route("/file/<int:fid>")
@login_required
def file_download(fid):
    if perm_level("view_contracts") == "none":
        abort(403)
    row = db.query_one(
        "SELECT file_name, file_path, contract_id FROM contract_files WHERE id=%s", (fid,))
    if not row or not os.path.exists(row["file_path"]):
        abort(404)
    audit("download", "contract", row["contract_id"], after={"file": row["file_name"]})
    return send_file(row["file_path"], as_attachment=True, download_name=row["file_name"])


@app.route("/contract/<int:cid>/history")
@login_required
def contract_history(cid):
    rows = db.query("""
        SELECT a.at, a.action, a.before, a.after, u.full_name
        FROM audit_log a LEFT JOIN users u ON u.id=a.user_id
        WHERE a.entity='contract' AND a.entity_id=%s
        ORDER BY a.at DESC LIMIT 100
    """, (cid,))
    c = db.query_one("SELECT number_text FROM contracts WHERE id=%s", (cid,))
    return render_template("history.html", rows=rows, c=c, cid=cid)


# ------------------------------------------------------------------
#  Связи между договорами
# ------------------------------------------------------------------

LINK_RU = {"reissue": "перевыпуск", "supplement": "допсоглашение",
           "annex": "приложение/смета", "same_object": "тот же объект"}
app.jinja_env.globals.update(LINK_RU=LINK_RU)


@app.route("/contract/<int:cid>/link", methods=["POST"])
@login_required
def link_add(cid):
    if not can("edit_contract"):
        abort(403)
    other = request.form.get("other_id")
    ltype = request.form.get("link_type")
    comment = (request.form.get("comment") or "").strip() or None
    if not (other and other.isdigit()) or ltype not in LINK_RU:
        flash("Выберите договор и тип связи")
        return redirect(url_for("contract", cid=cid) + "#links")
    other = int(other)
    if other == cid:
        flash("Нельзя связать договор с самим собой")
        return redirect(url_for("contract", cid=cid) + "#links")
    if not db.query_one("SELECT 1 FROM contracts WHERE id=%s", (other,)):
        abort(404)
    exists = db.query_one(
        "SELECT 1 FROM contract_links WHERE link_type=%s AND "
        "((parent_id=%s AND child_id=%s) OR (parent_id=%s AND child_id=%s))",
        (ltype, cid, other, other, cid))
    if not exists:
        db.execute("INSERT INTO contract_links(parent_id, child_id, link_type, comment) "
                   "VALUES(%s,%s,%s,%s)", (cid, other, ltype, comment))
        audit("link_add", "contract", cid, after={"other": other, "type": ltype})
    return redirect(url_for("contract", cid=cid) + "#links")


@app.route("/link/<int:lid>/delete", methods=["POST"])
@login_required
def link_delete(lid):
    if not can("edit_contract"):
        abort(403)
    row = db.query_one("SELECT parent_id FROM contract_links WHERE id=%s", (lid,))
    if not row:
        abort(404)
    db.execute("DELETE FROM contract_links WHERE id=%s", (lid,))
    audit("link_delete", "contract", row["parent_id"], after={"link_id": lid})
    return redirect(url_for("contract", cid=row["parent_id"]) + "#links")


# ------------------------------------------------------------------
#  Маршрут согласования (визирование)
# ------------------------------------------------------------------

def _can_manage_route():
    return can("approve_send") or can("edit_contract")


@app.route("/contract/<int:cid>/approval", methods=["POST"])
@login_required
def approval_add(cid):
    if not _can_manage_route():
        abort(403)
    if not db.query_one("SELECT 1 FROM contracts WHERE id=%s", (cid,)):
        abort(404)
    approver = request.form.get("approver_id")
    if not (approver and approver.isdigit()):
        flash("Выберите визирующего")
        return redirect(url_for("contract", cid=cid) + "#approval")
    nxt = db.query_one("SELECT coalesce(max(ord),0)+1 AS n FROM approvals WHERE contract_id=%s", (cid,))
    db.execute("INSERT INTO approvals(contract_id, ord, approver_id, created_by) "
               "VALUES(%s,%s,%s,%s)", (cid, nxt["n"], int(approver), g.user["id"]))
    audit("approval_add", "contract", cid, after={"approver_id": int(approver), "ord": nxt["n"]})
    return redirect(url_for("contract", cid=cid) + "#approval")


@app.route("/approval/<int:aid>/decide", methods=["POST"])
@login_required
def approval_decide(aid):
    a = db.query_one("SELECT id, contract_id, ord, approver_id, status FROM approvals WHERE id=%s", (aid,))
    if not a:
        abort(404)
    # решать может назначенный визирующий или у кого право «согласовать и отправить»
    if not (a["approver_id"] == g.user["id"] or can("approve_send")):
        abort(403)
    if a["status"] != "pending":
        flash("Этот шаг уже пройден")
        return redirect(url_for("contract", cid=a["contract_id"]) + "#approval")
    # только текущий (первый непройденный) шаг активен
    cur = db.query_one("SELECT id FROM approvals WHERE contract_id=%s AND status='pending' "
                       "ORDER BY ord, id LIMIT 1", (a["contract_id"],))
    if cur and cur["id"] != aid:
        flash("Сначала должен решить предыдущий согласующий")
        return redirect(url_for("contract", cid=a["contract_id"]) + "#approval")
    decision = request.form.get("decision")
    if decision not in ("approved", "rejected"):
        abort(400)
    comment = (request.form.get("comment") or "").strip() or None
    db.execute("UPDATE approvals SET status=%s, comment=%s, decided_at=now() WHERE id=%s",
               (decision, comment, aid))
    audit("approval_" + decision, "contract", a["contract_id"],
          after={"approval_id": aid, "comment": comment})
    flash("Согласовано" if decision == "approved" else "Отклонено с замечаниями")
    return redirect(url_for("contract", cid=a["contract_id"]) + "#approval")


@app.route("/approval/<int:aid>/delete", methods=["POST"])
@login_required
def approval_delete(aid):
    if not _can_manage_route():
        abort(403)
    a = db.query_one("SELECT id, contract_id FROM approvals WHERE id=%s", (aid,))
    if not a:
        abort(404)
    db.execute("DELETE FROM approvals WHERE id=%s", (aid,))
    audit("approval_delete", "contract", a["contract_id"], after={"approval_id": aid})
    return redirect(url_for("contract", cid=a["contract_id"]) + "#approval")


# ------------------------------------------------------------------
#  Сроки и напоминания
# ------------------------------------------------------------------

@app.route("/deadlines")
@login_required
def deadlines():
    show_pay = perm_level("payments") != "none"
    # неоплаченные платежи: просроченные и ближайшие (≤ 14 дней)
    pays = []
    if show_pay:
        pays = db.query("""
            SELECT p.id, p.kind, p.direction, p.planned_on, p.amount, p.condition,
                   (p.planned_on - current_date) AS days_left,
                   c.id AS cid, c.number_text, cp.name AS counterparty
            FROM payments p JOIN contracts c ON c.id=p.contract_id
            LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
            WHERE p.paid_on IS NULL AND p.planned_on IS NOT NULL
              AND p.planned_on <= current_date + 14
              AND c.stage NOT IN ('archived','cancelled')
            ORDER BY p.planned_on
        """)
    # гарантии на исходе (≤ 60 дней) и недавно истёкшие (последние 30 дней)
    warr = db.query("""
        SELECT c.id AS cid, c.number_text, c.warranty_until,
               (c.warranty_until - current_date) AS days_left,
               cp.name AS counterparty, o.address AS object_address
        FROM contracts c
        LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
        LEFT JOIN objects o ON o.id=c.object_id
        WHERE c.warranty_until IS NOT NULL
          AND c.warranty_until BETWEEN current_date - 30 AND current_date + 60
        ORDER BY c.warranty_until
    """)
    # срок действия договора истекает (≤ 30 дней) — важно для автопролонгации
    ending = db.query("""
        SELECT c.id AS cid, c.number_text, c.valid_to, c.auto_renewal,
               c.renewal_notice_days,
               (c.valid_to - current_date) AS days_left,
               cp.name AS counterparty
        FROM contracts c
        LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
        WHERE c.valid_to IS NOT NULL
          AND c.valid_to BETWEEN current_date AND current_date + 45
          AND c.stage NOT IN ('archived','cancelled')
        ORDER BY c.valid_to
    """)
    # этапы работ: не сделаны, плановая дата просрочена или близко (≤ 14 дней)
    stg = db.query("""
        SELECT s.id, s.name, s.volume, s.planned_on, s.amount,
               (s.planned_on - current_date) AS days_left,
               c.id AS cid, c.number_text, cp.name AS counterparty
        FROM stages s JOIN contracts c ON c.id=s.contract_id
        LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
        WHERE s.is_done = false AND s.planned_on IS NOT NULL
          AND s.planned_on <= current_date + 14
          AND c.stage NOT IN ('archived','cancelled')
        ORDER BY s.planned_on
    """)
    # задачи: незакрытые с подходящим или просроченным сроком (свои — приоритетно)
    tasks = []
    if can("tasks"):
        mine = "AND t.assignee_id = %s" if perm_level("tasks") == "own" else ""
        tparams = (g.user["id"],) if mine else ()
        tasks = db.query(f"""
            SELECT t.id, t.title, t.due_on, t.priority,
                   (t.due_on - current_date) AS days_left,
                   a.full_name AS assignee, c.id AS cid, c.number_text,
                   cp.name AS counterparty
            FROM tasks t
            LEFT JOIN users a ON a.id=t.assignee_id
            LEFT JOIN contracts c ON c.id=t.contract_id
            LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
            WHERE t.is_done = false AND t.due_on IS NOT NULL
              AND t.due_on <= current_date + 14 {mine}
            ORDER BY t.due_on
        """, tparams)
    return render_template("deadlines.html", pays=pays, warr=warr,
                           ending=ending, stg=stg, tasks=tasks, show_pay=show_pay)


def make_xlsx(sheet, headers, rows):
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = sheet[:31]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F5F73")
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    for r in rows:
        ws.append(r)
    from openpyxl.utils import get_column_letter
    for i, h in enumerate(headers, 1):
        maxlen = len(str(h))
        for r in rows:
            v = r[i - 1] if i - 1 < len(r) else ""
            maxlen = max(maxlen, len(str(v)) if v is not None else 0)
        ws.column_dimensions[get_column_letter(i)].width = min(48, max(10, maxlen + 2))
    bio = _io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@app.route("/reports")
@login_required
def reports():
    if not can("reports"):
        abort(403)
    return render_template("reports.html")


@app.route("/export/registry.xlsx")
@login_required
def export_registry():
    if not can("reports"):
        abort(403)
    rows = db.query("""
        SELECT c.number_text, c.signed_on, cp.name, ct.name AS type_name, o.address,
               c.amount, c.advance_pct, c.warranty_months, c.stage,
               (SELECT count(*) FROM review_findings rf WHERE rf.contract_id=c.id AND rf.resolution IS NULL) AS f
        FROM contracts c
        LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
        LEFT JOIN contract_types ct ON ct.code=c.type_code
        LEFT JOIN objects o ON o.id=c.object_id
        ORDER BY c.number_text
    """)
    data = [[r["number_text"], r["signed_on"], r["name"], r["type_name"], r["address"],
             float(r["amount"]) if r["amount"] else None,
             float(r["advance_pct"]) if r["advance_pct"] else None,
             r["warranty_months"], STAGE_RU.get(r["stage"], r["stage"]), r["f"]]
            for r in rows]
    bio = make_xlsx("Реестр договоров",
                    ["Номер", "Дата", "Заказчик", "Тип", "Объект", "Сумма",
                     "Аванс %", "Гарантия мес", "Стадия", "Замечаний"], data)
    audit("export", "report", 0, after={"report": "registry"})
    return send_file(bio, as_attachment=True, download_name="reestr.xlsx", mimetype=XLSX_MIME)


@app.route("/export/payments.xlsx")
@login_required
def export_payments():
    if perm_level("payments") == "none" or not can("reports"):
        abort(403)
    rows = db.query("""
        SELECT p.planned_on, c.number_text, cp.name, p.kind, p.direction,
               p.amount, p.paid_on, p.paid_amount, p.condition
        FROM payments p JOIN contracts c ON c.id=p.contract_id
        LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
        ORDER BY p.planned_on NULLS LAST
    """)
    data = [[r["planned_on"], r["number_text"], r["name"], PAY_KIND_RU.get(r["kind"], r["kind"]),
             DIR_RU.get(r["direction"], r["direction"]),
             float(r["amount"]) if r["amount"] else None, r["paid_on"],
             float(r["paid_amount"]) if r["paid_amount"] else None, r["condition"]]
            for r in rows]
    bio = make_xlsx("Платежи", ["Плановая дата", "Договор", "Заказчик", "Тип",
                                "Направление", "Сумма", "Оплачен", "Оплачено", "Условие"], data)
    audit("export", "report", 0, after={"report": "payments"})
    return send_file(bio, as_attachment=True, download_name="platezhi.xlsx", mimetype=XLSX_MIME)


@app.route("/export/warranty.xlsx")
@login_required
def export_warranty():
    if not can("reports"):
        abort(403)
    rows = db.query("""
        SELECT c.number_text, cp.name, o.address, c.warranty_months,
               c.commissioning_date, c.warranty_until, ct.name AS type_name
        FROM contracts c
        LEFT JOIN counterparties cp ON cp.id=c.counterparty_id
        LEFT JOIN objects o ON o.id=c.object_id
        LEFT JOIN contract_types ct ON ct.code=c.type_code
        WHERE c.warranty_months IS NOT NULL
        ORDER BY c.warranty_until NULLS LAST, c.number_text
    """)
    data = [[r["number_text"], r["name"], r["address"], r["warranty_months"],
             r["commissioning_date"], r["warranty_until"], r["type_name"]] for r in rows]
    bio = make_xlsx("Гарантии", ["Договор", "Заказчик", "Объект", "Гарантия мес",
                                 "Дата ввода", "Действует до", "Тип"], data)
    audit("export", "report", 0, after={"report": "warranty"})
    return send_file(bio, as_attachment=True, download_name="garantii.xlsx", mimetype=XLSX_MIME)


@app.route("/health")
def health():
    try:
        n = db.query_one("SELECT count(*) AS n FROM contracts")
        return {"ok": True, "contracts": n["n"]}
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error": str(e)}, 500


@app.route("/admin/reload")
def admin_reload():
    """Перезапуск процесса для подхвата нового кода. Батник поднимет заново."""
    if request.args.get("token") != os.getenv("RELOAD_TOKEN"):
        abort(403)
    threading.Timer(0.4, lambda: os._exit(0)).start()
    return "reloading"


@app.errorhandler(403)
def forbidden(e):
    return render_template("error.html", code=403, msg="Нет доступа"), 403


@app.errorhandler(404)
def notfound(e):
    return render_template("error.html", code=404, msg="Не найдено"), 404


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=8080, debug=True)
